#!/usr/bin/env python3
"""下载飞书导出表里的原曲、人声参考、伴奏和干声。"""

from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import re
import sys
from dataclasses import asdict, dataclass
from email.message import Message
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen


ROOT = Path(__file__).resolve().parent.parent
# 和 render_feishu_mix_compare_batch.py 的默认读取目录保持一致，避免下载后批处理找不到。
DEFAULT_OUTPUT_DIR = ROOT.parent / "feishu_long_audio_screened"
DEFAULT_SHEET_NAME = "长音频-筛选"
DRY_COL = 27
ORIGINAL_COL = 4
REFERENCE_VOCAL_COL = 7
ACCOMP_COL = 8
CASE_COL = 1
NAME_COL = 2


@dataclass
class DownloadRecord:
    row: int
    role: str
    case_name: str
    extra_name: str
    url: str
    output_path: str
    status: str
    error: str | None = None


@dataclass
class SheetRecord:
    row: int
    case_name: str
    extra_name: str
    original_urls: list[str]
    reference_vocal_urls: list[str]
    accompaniment_urls: list[str]
    dry_urls: list[str]


def sanitize_filename(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value.strip())
    value = re.sub(r"\s+", " ", value)
    value = value.strip(" ._")
    return value or "unnamed"


def split_urls(value: str) -> list[str]:
    if not value:
        return []
    urls = re.findall(r"https?://[^\s,，;；)）\"']+", value)
    return [url.rstrip(".") for url in urls]


def formula_urls(value: object) -> list[str]:
    if not isinstance(value, str) or not value.startswith("="):
        return []
    return split_urls(value)


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_rows_from_xlsx(path: Path, sheet_name: str) -> list[tuple[int, list[str], dict[int, list[str]]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required for .xlsx files. Install it with: python -m pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise SystemExit(f"Sheet not found: {sheet_name}. Available sheets: {available}")
    ws = wb[sheet_name]

    rows: list[tuple[int, list[str], dict[int, list[str]]]] = []
    for row in ws.iter_rows():
        values = [cell_text(cell.value) for cell in row]
        links: dict[int, list[str]] = {}
        for idx, cell in enumerate(row, start=1):
            found: list[str] = []
            if cell.hyperlink and cell.hyperlink.target:
                found.append(str(cell.hyperlink.target))
            found.extend(formula_urls(cell.value))
            found.extend(split_urls(cell_text(cell.value)))
            if found:
                links[idx] = list(dict.fromkeys(found))
        rows.append((row[0].row if row else len(rows) + 1, values, links))
    return rows


def load_rows_from_csv(path: Path, encoding: str) -> list[tuple[int, list[str], dict[int, list[str]]]]:
    text = path.read_text(encoding=encoding)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    rows: list[tuple[int, list[str], dict[int, list[str]]]] = []
    for row_num, row in enumerate(csv.reader(text.splitlines(), dialect), start=1):
        values = [cell_text(value) for value in row]
        links = {
            idx: split_urls(value)
            for idx, value in enumerate(values, start=1)
            if split_urls(value)
        }
        rows.append((row_num, values, links))
    return rows


def load_rows(path: Path, sheet_name: str, encoding: str) -> list[tuple[int, list[str], dict[int, list[str]]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return load_rows_from_xlsx(path, sheet_name)
    if suffix == ".csv":
        return load_rows_from_csv(path, encoding)
    raise SystemExit(f"Unsupported file type: {path.suffix}. Please export as .xlsx or .csv.")


def col_value(values: list[str], index: int) -> str:
    return values[index - 1].strip() if len(values) >= index else ""


def col_urls(values: list[str], links: dict[int, list[str]], index: int) -> list[str]:
    urls = links.get(index, [])
    if not urls:
        urls = split_urls(col_value(values, index))
    return list(dict.fromkeys(urls))


def output_stem(case_name: str, extra_name: str, role: str, row: int | None = None) -> str:
    base = sanitize_filename(f"{case_name}{extra_name}" if extra_name else case_name)
    stem = f"{base}_{role}"
    # “线上数据-5.25”在同一个 case name 下有多行不同音频，文件名必须带行号。
    if case_name == "线上数据-5.25" and row is not None:
        stem = f"{stem}_row{row}"
    return stem


def filename_from_headers(headers: Message, url: str, default_stem: str, fallback_ext: str) -> str:
    disposition = headers.get("Content-Disposition", "")
    match = re.search(r"filename\\*=UTF-8''([^;]+)", disposition, flags=re.I)
    if match:
        ext = Path(unquote(match.group(1))).suffix
        if ext:
            return default_stem + ext
    match = re.search(r'filename="?([^";]+)"?', disposition, flags=re.I)
    if match:
        ext = Path(unquote(match.group(1))).suffix
        if ext:
            return default_stem + ext

    path_ext = Path(unquote(urlparse(url).path)).suffix
    if path_ext:
        return default_stem + path_ext

    content_type = headers.get_content_type() if hasattr(headers, "get_content_type") else headers.get("Content-Type", "")
    guessed = mimetypes.guess_extension(content_type.split(";")[0].strip()) if content_type else None
    return default_stem + (guessed or fallback_ext)


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    index = 2
    while True:
        candidate = path.with_name(f"{stem}_{index:02d}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def download_url(url: str, output_dir: Path, stem: str, headers: dict[str, str], overwrite: bool) -> Path:
    request = Request(url, headers=headers)
    with urlopen(request, timeout=90) as response:
        filename = filename_from_headers(response.headers, response.url, stem, ".wav")
        output_path = output_dir / sanitize_filename(filename)
        if not overwrite:
            output_path = unique_path(output_path)
        with output_path.open("wb") as handle:
            while True:
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                handle.write(chunk)
    return output_path


def build_headers(cookie: str | None, user_agent: str) -> dict[str, str]:
    headers = {"User-Agent": user_agent}
    if cookie:
        headers["Cookie"] = cookie
    return headers


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("sheet_file", type=Path, help="飞书导出的 .xlsx/.csv。")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--encoding", default="utf-8-sig")
    parser.add_argument("--cookie", help="Optional Cookie header if attachment links require Feishu login.")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--user-agent", default="Mozilla/5.0")
    args = parser.parse_args()

    rows = load_rows(args.sheet_file, args.sheet_name, args.encoding)
    dry_dir = args.out_dir / "干声"
    accomp_dir = args.out_dir / "伴奏"
    dry_dir.mkdir(parents=True, exist_ok=True)
    accomp_dir.mkdir(parents=True, exist_ok=True)

    headers = build_headers(args.cookie, args.user_agent)
    records: list[DownloadRecord] = []
    sheet_records: list[SheetRecord] = []

    for row_num, values, links in rows:
        if row_num < args.start_row:
            continue
        case_name = col_value(values, CASE_COL)
        if not case_name:
            continue
        extra_name = col_value(values, NAME_COL)
        original_urls = col_urls(values, links, ORIGINAL_COL)
        reference_vocal_urls = col_urls(values, links, REFERENCE_VOCAL_COL)
        accompaniment_urls = col_urls(values, links, ACCOMP_COL)
        dry_urls = col_urls(values, links, DRY_COL)
        # 保存每行的 D/G/H/AA 列 URL，后续批处理可以准确带上参考曲。
        sheet_records.append(SheetRecord(
            row=row_num,
            case_name=case_name,
            extra_name=extra_name,
            original_urls=original_urls,
            reference_vocal_urls=reference_vocal_urls,
            accompaniment_urls=accompaniment_urls,
            dry_urls=dry_urls,
        ))
        jobs = [
            ("原曲", args.out_dir / "原曲", original_urls),
            ("原曲人声", args.out_dir / "原曲人声", reference_vocal_urls),
            ("干声", dry_dir, dry_urls),
            ("伴奏", accomp_dir, accompaniment_urls),
        ]
        for role, role_dir, urls in jobs:
            role_dir.mkdir(parents=True, exist_ok=True)
            for url_index, url in enumerate(urls, start=1):
                stem = output_stem(case_name, extra_name, role, row=row_num)
                if len(urls) > 1:
                    stem = f"{stem}_{url_index:02d}"
                output_path = role_dir / f"{stem}.wav"
                if args.dry_run:
                    records.append(DownloadRecord(row_num, role, case_name, extra_name, url, str(output_path), "dry_run"))
                    continue
                try:
                    saved = download_url(url, role_dir, stem, headers, args.overwrite)
                    records.append(DownloadRecord(row_num, role, case_name, extra_name, url, str(saved), "ok"))
                    print(f"[ok] row {row_num} {role}: {saved}")
                except (HTTPError, URLError, OSError) as exc:
                    records.append(DownloadRecord(row_num, role, case_name, extra_name, url, str(output_path), "failed", str(exc)))
                    print(f"[failed] row {row_num} {role}: {exc}", file=sys.stderr)

    manifest = args.out_dir / "download_manifest.json"
    manifest.write_text(
        json.dumps([asdict(record) for record in records], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    # 批处理脚本读取这个聚合文件来决定每行输入和参考素材。
    sheet_records_path = args.out_dir / "sheet_records.json"
    sheet_records_path.write_text(
        json.dumps({"records": [asdict(record) for record in sheet_records]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    ok_count = sum(1 for record in records if record.status == "ok")
    failed_count = sum(1 for record in records if record.status == "failed")
    print(json.dumps({
        "manifest": str(manifest),
        "sheet_records": str(sheet_records_path),
        "records": len(records),
        "ok": ok_count,
        "failed": failed_count,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
